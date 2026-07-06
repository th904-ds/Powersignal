"""
전처리 파이프라인

Steps:
  1. 변수 제거 (상수·기상 원시)
  2. gen 5종 결측 보간 → 일합산 24h 반복 → ratio 재계산
  3. 미소결측 처리 (gen 3종 0.1%, smp_decision_cnt_LNG 0.2%)
  3b. SMP 이상치 처리
  W.  기상변수 이상치·결측 처리 + avg_temp_c_sq 생성 + 병합
  4. datetime 피처
  5. 공휴일 피처
  6. 수요예측 래그/차분 (공통)
  7. SMP 자기회귀 피처 (Model 1 전용)
  8. 훈련 기간 필터 & 저장
"""
from __future__ import annotations

import pandas as pd
import numpy as np
from pathlib import Path

# ── 경로 ──────────────────────────────────────────────────────────────
# GitHub Actions/Render/로컬에서 모두 동작하도록 현재 파일 위치 기준으로 프로젝트 루트를 계산한다.
_PROJECT = Path(__file__).resolve().parents[2]
FEAT      = _PROJECT / "00.collector" / "data" / "features"
MANUAL    = _PROJECT / "00.collector" / "data" / "manual"
OUT       = _PROJECT / "01.preprocessing" / "output"
# sukub 5분 데이터 소스 ①: power_supply_today 일별 스냅샷 parquet
_SUKUB_TODAY_DIR  = _PROJECT / "00.collector" / "data" / "processed" / "power_supply_today"
# sukub 5분 데이터 소스 ②: sukub.do 에서 수동 다운로드한 CSV 파일 모음
_SUKUB_MANUAL_DIR = MANUAL / "sukub"
OUT.mkdir(parents=True, exist_ok=True)

TRAIN_START = "2023-01-01"
TRAIN_END   = None

# ── 상수 ──────────────────────────────────────────────────────────────
WEATHER_COLS       = ["temp_c", "humidity_pct", "wind_speed_ms", "dew_point_c"]
CONST_COLS         = ["fuel_cost_무연탄", "fuel_cost_원자력"]   # EDA4: unique=1
STALE_SUPPLY_COLS  = ["daily_max_demand", "daily_min_demand", "reserve_to_max_demand"]

GEN_INTERP    = ["gen_LNG", "gen_원자력", "gen_유연탄", "gen_무연탄", "gen_유전"]
GEN_MINOR     = ["gen_신재생·기타", "gen_수력", "gen_양수"]   # 0.1% 결측
GEN_ALL       = GEN_INTERP + GEN_MINOR
GEN_RATIO_ALL = [c + "_ratio" for c in GEN_ALL]

SMP_LAG_HOURS    = [1, 24, 48, 72, 168, 336]
SMP_ROLL_WINS    = [24, 168]
LOAD_LAG_HOURS   = [24, 168]
RESERVE_LAG_HOURS = [24, 168]
RESERVE_ROLL_WINS = [24, 168]

# sukub 원시 컬럼명 → 영문 매핑
# ① 한글 (sukub.do 수동 CSV 다운로드)
# ② 영문 camelCase (getSukub5mToday API 수집 parquet)
_SUKUB_RENAME = {
    "기준일시":         "datetime",
    "공급능력(MW)":     "supply_capacity",
    "현재수요(MW)":     "current_demand",
    "최대예측수요(MW)": "forecast_load",
    "공급예비력(MW)":   "supply_reserve_power",
    "공급예비율(%)":    "supply_reserve_rate",
    "운영예비력(MW)":   "operating_reserve_power",
    "운영예비율(%)":    "operating_reserve_rate",
    # API parquet 컬럼명 (power_supply_today)
    "baseDatetime":     "datetime",
    "suppAbility":      "supply_capacity",
    "currPwrTot":       "current_demand",
    "forecastLoad":     "forecast_load",
    "suppReservePwr":   "supply_reserve_power",
    "suppReserveRate":  "supply_reserve_rate",
    "operReservePwr":   "operating_reserve_power",
    "operReserveRate":  "operating_reserve_rate",
}

# 5분→1시간 집계 방식
_SUKUB_AGG = {
    "supply_capacity":          "min",
    "current_demand":           "max",
    "forecast_load":            "max",
    "supply_reserve_power":     "min",
    "supply_reserve_rate":      "min",
    "operating_reserve_power":  "min",
    "operating_reserve_rate":   "min",
}


# ══════════════════════════════════════════════════════════════════════
# STEP 1 — 변수 제거
# ══════════════════════════════════════════════════════════════════════
def step1_drop(df: pd.DataFrame) -> pd.DataFrame:
    drop = [c for c in CONST_COLS + WEATHER_COLS + STALE_SUPPLY_COLS if c in df.columns]
    df = df.drop(columns=drop)
    print(f"[Step 1] 제거: {drop}  →  {df.shape[1]}개 컬럼 남음")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 2 — gen 5종 결측 보간 → 일합산 24h 반복 → ratio 재계산
# ══════════════════════════════════════════════════════════════════════
def step2_gen_impute(df: pd.DataFrame) -> pd.DataFrame:
    """
    1) 시간 선형 보간 (hourly level, 경계 ffill/bfill로 가장자리 처리)
    2) 일합산(daily sum) 후 하루 24시간 반복 (모델팀 방식)
    3) gen_total 재계산
    4) gen_*_ratio 재계산
    """
    df = df.copy()

    # ── 2-1. hourly 선형 보간 ──────────────────────────────────────────
    # method="time"은 DatetimeIndex 필요 → 임시 설정
    df = df.set_index("datetime")
    for col in GEN_INTERP:
        if col not in df.columns:
            continue
        before = df[col].isnull().sum()
        df[col] = df[col].interpolate(method="time").ffill().bfill()
        after = df[col].isnull().sum()
        print(f"  [2-1] {col}: {before:,} → {after} 결측")
    df = df.reset_index()

    # ── 2-2. 일합산 후 24h 반복 ───────────────────────────────────────
    # datetime 인덱스 기준으로 _date 컬럼 생성
    df["_date"] = df["datetime"].dt.normalize()

    for col in GEN_ALL:
        if col not in df.columns:
            continue
        # 일합산
        daily_sum = df.groupby("_date")[col].transform("sum")
        df[col] = daily_sum   # 24시간 동일값으로 덮어쓰기

    print(f"  [2-2] gen {len(GEN_ALL)}종 → 일합산 24h 반복 완료")

    # ── 2-3. gen_total 재계산 ─────────────────────────────────────────
    gen_cols_present = [c for c in GEN_ALL if c in df.columns]
    df["gen_total"] = df[gen_cols_present].sum(axis=1)
    print(f"  [2-3] gen_total 재계산 완료")

    # ── 2-4. ratio 재계산 ─────────────────────────────────────────────
    for col in gen_cols_present:
        ratio_col = col + "_ratio"
        df[ratio_col] = df[col] / df["gen_total"].replace(0, np.nan)
    print(f"  [2-4] gen_*_ratio {len(gen_cols_present)}종 재계산 완료")

    df = df.drop(columns=["_date"])

    # 검증
    remaining = df[[c for c in GEN_INTERP if c in df.columns]].isnull().sum().sum()
    print(f"  [검증] gen 5종 잔여 결측: {remaining}")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 3 — 미소결측 처리
# ══════════════════════════════════════════════════════════════════════
def step3_minor_impute(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # gen 3종 (신재생·기타, 수력, 양수) + total: 선형 보간
    minor_gen = [c for c in GEN_MINOR + ["gen_total"] + GEN_RATIO_ALL if c in df.columns]
    df = df.set_index("datetime")
    for col in minor_gen:
        before = df[col].isnull().sum()
        if before > 0:
            df[col] = df[col].interpolate(method="time").ffill().bfill()
            print(f"  [3] {col}: {before} → 0 결측 (선형보간)")
    df = df.reset_index()

    # smp_decision_cnt_LNG: ffill
    col = "smp_decision_cnt_LNG"
    if col in df.columns:
        before = df[col].isnull().sum()
        if before > 0:
            df[col] = df[col].ffill().bfill()
            print(f"  [3] {col}: {before} → {df[col].isnull().sum()} 결측 (ffill)")

    remaining = df.drop(columns=["datetime"], errors="ignore").isnull().sum().sum()
    print(f"  [검증] 전체 잔여 결측: {remaining}")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 3.5 — SMP 이상치 처리
# ══════════════════════════════════════════════════════════════════════
def step3b_smp_outlier(df: pd.DataFrame) -> pd.DataFrame:
    """
    SMP = 0 (25건) → ffill (직전 시간 SMP로 대체)
    SMP 고가 spike (>272원, 132건) → 유지 (실제 에너지위기 현상, clip 안 함)
    """
    df = df.copy()
    n_zero = (df["smp"] == 0).sum()
    if n_zero > 0:
        df["smp"] = df["smp"].replace(0, np.nan).ffill()
        print(f"  [3b] SMP=0 {n_zero}건 → ffill 처리")
    spike_thr = df["smp"].quantile(0.995)
    n_spike = (df["smp"] >= spike_thr).sum()
    print(f"  [3b] SMP spike(≥{spike_thr:.1f}원) {n_spike}건 → 유지 (실제 시장 데이터)")
    print(f"  [3b] SMP 범위: {df['smp'].min():.2f} ~ {df['smp'].max():.2f}")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP W — 기상변수 처리 + 병합
# ══════════════════════════════════════════════════════════════════════
_WX_PATH = _PROJECT / "01.preprocessing" / "output" / "weather_hourly.parquet"

_ZERO_FILL_WX = ["avg_precip_mm", "avg_snow_cm", "avg_sunshine_hr", "avg_solar_MJm2"]
_PHYS_BOUNDS  = {
    "avg_temp_c":        (-22,  42),
    "avg_humidity_pct":  (  0, 100),
    "avg_wind_speed_ms": (  0,  30),
    "avg_precip_mm":     (  0,  80),
    "avg_snow_cm":       (  0,  25),
    "avg_sunshine_hr":   (  0,   1),
    "avg_solar_MJm2":    (  0,   4),
}
WX_COLS = [
    "avg_temp_c", "avg_humidity_pct", "avg_wind_speed_ms",
    "avg_precip_mm", "avg_snow_cm", "avg_sunshine_hr", "avg_solar_MJm2",
    "avg_temp_c_sq",
]


def stepW_weather_merge(df: pd.DataFrame) -> pd.DataFrame:
    """
    1) weather_hourly.parquet 로드 (avg_ 7개 컬럼)
    2) 0-fill: 강수·적설·일조·일사
    3) 물리적 클리핑
    4) avg_temp_c_sq 생성
    5) datetime 기준으로 df 에 병합
    """
    if not _WX_PATH.exists():
        print(f"[Step W] 경고: {_WX_PATH} 없음 — 기상 스킵")
        return df

    wx = pd.read_parquet(_WX_PATH)
    wx["datetime"] = pd.to_datetime(wx["datetime"])
    avg_cols = [c for c in wx.columns if c.startswith("avg_")]
    wx = wx[["datetime"] + avg_cols].copy()

    # 0-fill
    for c in _ZERO_FILL_WX:
        if c in wx.columns:
            wx[c] = wx[c].fillna(0)

    # linear 보간 (기온·습도·풍속, 거의 없음)
    other = [c for c in avg_cols if c not in _ZERO_FILL_WX]
    for c in other:
        if wx[c].isnull().any():
            wx[c] = wx[c].interpolate(method="linear").ffill().bfill()

    # 물리적 클리핑
    for c, (lo, hi) in _PHYS_BOUNDS.items():
        if c in wx.columns:
            wx[c] = wx[c].clip(lower=lo, upper=hi)

    # 파생변수
    wx["avg_temp_c_sq"] = wx["avg_temp_c"] ** 2

    # 병합
    df = df.merge(wx[["datetime"] + WX_COLS], on="datetime", how="left")
    wx_null = df[WX_COLS].isnull().sum().sum()
    print(f"[Step W] 기상 {len(WX_COLS)}개 컬럼 병합 완료  결측={wx_null}  "
          f"shape={df.shape}")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP NBTP — 수요자원거래시장 NBTP 월별 전파
# ══════════════════════════════════════════════════════════════════════
_NBTP_PATH = MANUAL / "2026년도 3월 수요자원거래시장 현황 상세.xlsx"
_NBTP_ACCEPTED_ROWS = [5, 12, 19, 26, 33]


def stepNBTP_merge(df: pd.DataFrame) -> pd.DataFrame:
    import openpyxl

    if not _NBTP_PATH.exists():
        print(f"[Step NBTP] 경고: {_NBTP_PATH} 없음 — 스킵")
        return df

    wb = openpyxl.load_workbook(_NBTP_PATH)
    ws = wb[wb.sheetnames[3]]  # 낙찰결과표 현황

    records = []
    for nbtp_row in _NBTP_ACCEPTED_ROWS:
        for col in range(3, 15):  # C~N = 1월~12월
            year  = ws.cell(row=nbtp_row - 2, column=col).value
            month = ws.cell(row=nbtp_row - 1, column=col).value
            if not isinstance(year, (int, float)):
                continue
            v_acc = ws.cell(row=nbtp_row,     column=col).value
            v_bid = ws.cell(row=nbtp_row + 1, column=col).value
            records.append({
                "_month":        pd.Timestamp(int(year), int(month), 1),
                "nbtp_accepted": v_acc if v_acc not in (None, "-") else None,
                "nbtp_bid":      v_bid if v_bid not in (None, "-") else None,
            })

    nbtp = pd.DataFrame(records).dropna(subset=["nbtp_accepted"])

    df = df.copy()
    df["_month"] = df["datetime"].dt.to_period("M").dt.to_timestamp()
    df = df.merge(nbtp, on="_month", how="left")
    df = df.drop(columns=["_month"])

    n_acc = df["nbtp_accepted"].isnull().sum()
    n_bid = df["nbtp_bid"].isnull().sum()
    print(f"[Step NBTP] nbtp_accepted 결측={n_acc}, nbtp_bid 결측={n_bid}  shape={df.shape}")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP SUKUB — 5분단위 전력수급현황 집계 & 병합
# ══════════════════════════════════════════════════════════════════════
def _normalize_sukub(df: pd.DataFrame) -> pd.DataFrame:
    """컬럼명을 영문 스네이크케이스로 정규화하고 불필요한 컬럼 제거."""
    df = df[[c for c in df.columns if not c.startswith("_")]]
    df = df.rename(columns={k: v for k, v in _SUKUB_RENAME.items() if k in df.columns})
    # 집계에 필요한 컬럼만 유지
    keep = {"datetime"} | set(_SUKUB_AGG.keys())
    df = df[[c for c in df.columns if c in keep]]
    return df


def _load_sukub_parts() -> list[pd.DataFrame]:
    """sukub 5분 데이터를 두 소스에서 로드 (각 파트를 컬럼 정규화 후 반환).

    소스 ①: power_supply_today parquet (API, 영문 camelCase, 매일 스냅샷)
    소스 ②: 00.collector/data/manual/sukub/*.csv    (sukub.do 수동 다운로드, 한글 컬럼, EUC-KR)
    """
    parts = []

    # ① API parquet
    if _SUKUB_TODAY_DIR.exists():
        for f in sorted(_SUKUB_TODAY_DIR.glob("*.parquet")):
            try:
                tmp = _normalize_sukub(pd.read_parquet(f))
                parts.append(tmp)
            except Exception as e:
                print(f"  [SUKUB] parquet 로드 실패 {f.name}: {e}")

    # ② 수동 CSV (한글 컬럼, EUC-KR 인코딩)
    if _SUKUB_MANUAL_DIR.exists():
        for f in sorted(_SUKUB_MANUAL_DIR.glob("*.csv")):
            try:
                tmp = _normalize_sukub(pd.read_csv(f, encoding="euc-kr"))
                parts.append(tmp)
            except Exception as e:
                print(f"  [SUKUB] CSV 로드 실패 {f.name}: {e}")

    return parts


def stepSUKUB_merge(df: pd.DataFrame) -> pd.DataFrame:
    """
    power_supply_today parquet + 수동 CSV를 통합하여
    5분→1시간으로 집계한 뒤 df에 병합.

    - supply_capacity / supply_reserve_power / supply_reserve_rate:
      기존 일별 컬럼을 drop하고 시간별 집계값으로 교체
    - current_demand / forecast_load / operating_reserve_power / operating_reserve_rate:
      신규 추가
    """
    parts = _load_sukub_parts()
    if not parts:
        print("[Step SUKUB] 경고: sukub 데이터 없음 (power_supply_today parquet 또는 "
              "data/manual/sukub/*.csv 필요) — 스킵")
        return df

    # 각 파트는 이미 _normalize_sukub 에서 컬럼 정규화 완료
    raw = pd.concat(parts, ignore_index=True)

    # datetime 파싱 (YYYYMMDDHHMMSS 문자열 또는 Timestamp)
    if raw["datetime"].dtype == object or str(raw["datetime"].dtype) == "string":
        raw["datetime"] = pd.to_datetime(raw["datetime"].astype(str), format="%Y%m%d%H%M%S")
    else:
        raw["datetime"] = pd.to_datetime(raw["datetime"])

    # 수치 컬럼 강제 변환 (CSV에서 문자열로 들어오는 경우)
    for col in _SUKUB_AGG:
        if col in raw.columns:
            raw[col] = pd.to_numeric(raw[col], errors="coerce")

    # 5분 → 1시간 집계
    raw["datetime"] = raw["datetime"].dt.floor("h")
    agg_map = {c: m for c, m in _SUKUB_AGG.items() if c in raw.columns}
    hourly = raw.groupby("datetime").agg(agg_map).reset_index()
    hourly = hourly.drop_duplicates(subset=["datetime"])

    # 기존 일별 공급 컬럼 제거 (교체 대상)
    replace_old = ["supply_capacity", "supply_reserve_power", "supply_reserve_rate"]
    df = df.drop(columns=[c for c in replace_old if c in df.columns])

    # 시간별 집계값 병합
    df = df.merge(hourly, on="datetime", how="left")

    null_before = {c: int(df[c].isnull().sum()) for c in _SUKUB_AGG if c in df.columns}

    # 결측 보간: 시간 선형 보간 → 엣지 ffill/bfill
    # 에너지 수급값은 연속적으로 변화하므로 시간 기반 선형 보간이 적합
    sukub_cols = [c for c in _SUKUB_AGG if c in df.columns]
    df = df.set_index("datetime")
    for col in sukub_cols:
        if df[col].isnull().any():
            df[col] = df[col].interpolate(method="time").ffill().bfill()
    df = df.reset_index()

    null_after = {c: int(df[c].isnull().sum()) for c in sukub_cols}
    src_info = (f"today_parquet={len(list(_SUKUB_TODAY_DIR.glob('*.parquet'))) if _SUKUB_TODAY_DIR.exists() else 0}, "
                f"manual_csv={len(list(_SUKUB_MANUAL_DIR.glob('*.csv'))) if _SUKUB_MANUAL_DIR.exists() else 0}")
    print(f"[Step SUKUB] 집계 {len(hourly):,}행 병합 완료  ({src_info})  shape={df.shape}")
    print(f"  병합 후 결측: {null_before}")
    print(f"  보간 후 결측: {null_after}")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP RESERVE LAGS — 예비력 lag/rolling 피처
# ══════════════════════════════════════════════════════════════════════
_RESERVE_BASE_COLS = [
    "operating_reserve_rate",
    "operating_reserve_power",
    "supply_capacity",
    "current_demand",
    "forecast_load",
]


def stepRESERVE_lags(df: pd.DataFrame) -> pd.DataFrame:
    """
    sukub 집계 컬럼들의 lag/rolling 피처 생성.

    lag:
      - operating_reserve_rate / _power / supply_capacity / current_demand / forecast_load
        × lag24 / lag168

    rolling (operating_reserve_rate만, lag24 기준):
      - mean / min / std  × window 24h / 168h

    ratio:
      - operating_reserve_power_to_demand_lag24
        = operating_reserve_power_lag24 / current_demand_lag24
    """
    df = df.copy()
    n_new = 0

    # lag 피처
    for col in _RESERVE_BASE_COLS:
        if col not in df.columns:
            continue
        for lag in RESERVE_LAG_HOURS:
            df[f"{col}_lag{lag}"] = df[col].shift(lag)
            n_new += 1

    # rolling 피처 (운영예비율)
    rate_col = "operating_reserve_rate"
    if rate_col in df.columns:
        for win in RESERVE_ROLL_WINS:
            base = df[rate_col].shift(24)
            rolled = base.rolling(window=win, min_periods=win // 2)
            df[f"{rate_col}_roll_mean_{win}_lag24"] = rolled.mean()
            df[f"{rate_col}_roll_min_{win}_lag24"]  = rolled.min()
            df[f"{rate_col}_roll_std_{win}_lag24"]  = rolled.std()
            n_new += 3

    # 비율 파생 피처
    if "operating_reserve_power_lag24" in df.columns and "current_demand_lag24" in df.columns:
        df["operating_reserve_power_to_demand_lag24"] = (
            df["operating_reserve_power_lag24"] / df["current_demand_lag24"].replace(0, np.nan)
        )
        n_new += 1

    print(f"[Step RESERVE] 예비력 lag/rolling 피처 {n_new}개 생성 완료")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 4 — datetime 피처
# ══════════════════════════════════════════════════════════════════════
def step4_datetime_feats(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    dt = df["datetime"]

    df["hour_of_day"] = dt.dt.hour.replace(0, 24)   # 00시 → 24시 (컨벤션)
    df["weekday"]     = dt.dt.dayofweek              # 0=월 ~ 6=일
    df["month_num"]   = dt.dt.month
    df["is_weekend"]  = (df["weekday"] >= 5).astype(int)
    df["date_key"]    = dt.dt.date.astype(str)
    df["month_key"]   = dt.dt.to_period("M").astype(str)

    # 순환 인코딩
    df["hour_sin"]  = np.sin(2 * np.pi * df["hour_of_day"] / 24)
    df["hour_cos"]  = np.cos(2 * np.pi * df["hour_of_day"] / 24)
    df["month_sin"] = np.sin(2 * np.pi * df["month_num"] / 12)
    df["month_cos"] = np.cos(2 * np.pi * df["month_num"] / 12)

    print(f"[Step 4] datetime 피처 9종 생성 완료")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 5 — 공휴일 피처
# ══════════════════════════════════════════════════════════════════════
def step5_holiday_feats(df: pd.DataFrame) -> pd.DataFrame:
    try:
        import holidays
    except ImportError:
        print("[Step 5] holidays 패키지 없음. pip install holidays")
        df["is_holiday"] = 0
        df["is_before_holiday"] = 0
        df["is_after_holiday"] = 0
        return df

    df = df.copy()
    kr_holidays = holidays.KR(years=range(2022, 2027))

    from datetime import timedelta
    holiday_set = set(kr_holidays.keys())   # datetime.date 집합
    dates_d = df["datetime"].dt.date        # Series[datetime.date]
    df["is_holiday"]        = dates_d.map(lambda d: int(d in holiday_set))
    df["is_before_holiday"] = dates_d.map(lambda d: int(d + timedelta(days=1) in holiday_set))
    df["is_after_holiday"]  = dates_d.map(lambda d: int(d - timedelta(days=1) in holiday_set))

    n_holiday = df["is_holiday"].sum()
    print(f"[Step 5] 공휴일 피처 생성 완료  (공휴일 시간수: {n_holiday:,})")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 6 — 수요예측 래그/차분 (공통)
# ══════════════════════════════════════════════════════════════════════
def step6_load_lags(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for base_col in ["jlfd", "slfd", "mlfd"]:
        if base_col not in df.columns:
            continue
        for lag in LOAD_LAG_HOURS:
            df[f"{base_col}_lag{lag}"] = df[base_col].shift(lag)
        df[f"{base_col}_diff_24"]       = df[base_col] - df[base_col].shift(24)
        df[f"{base_col}_pct_change_24"] = df[f"{base_col}_diff_24"] / df[base_col].shift(24)

    n_new = 3 * (len(LOAD_LAG_HOURS) + 2)
    print(f"[Step 6] 수요예측 래그/차분 {n_new}개 생성 완료")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 7 — SMP 자기회귀 피처 (Model 1 전용)
# ══════════════════════════════════════════════════════════════════════
def step7_smp_lags(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    smp = df["smp"]

    # 래그
    for lag in SMP_LAG_HOURS:
        df[f"smp_lag{lag}"] = smp.shift(lag)

    # 롤링
    for win in SMP_ROLL_WINS:
        rolled = smp.shift(1).rolling(window=win, min_periods=win // 2)
        df[f"smp_roll_mean_{win}"] = rolled.mean()
        df[f"smp_roll_std_{win}"]  = rolled.std()
        df[f"smp_roll_max_{win}"]  = rolled.max()
        df[f"smp_roll_min_{win}"]  = rolled.min()

    n_new = len(SMP_LAG_HOURS) + len(SMP_ROLL_WINS) * 4
    print(f"[Step 7] SMP 자기회귀 피처 {n_new}개 생성 완료")
    return df


# ══════════════════════════════════════════════════════════════════════
# STEP 8 — 훈련 기간 필터 & 저장
# ══════════════════════════════════════════════════════════════════════
def step8_filter_save(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if TRAIN_END is None:
        train = df[df["datetime"] >= TRAIN_START].copy()
    else:
        train = df[(df["datetime"] >= TRAIN_START) &
                   (df["datetime"] <= TRAIN_END)].copy()

    print(f"[Step 8] model_features 기간 필터: {len(train):,}행  "
          f"({train['datetime'].min()} ~ {train['datetime'].max()})")

    m1_drop = ["date_key", "month_key"]
    m1 = train.drop(columns=[c for c in m1_drop if c in train.columns])

    smp_ar_cols = (
        [f"smp_lag{l}" for l in SMP_LAG_HOURS] +
        [f"smp_roll_{s}_{w}" for s in ["mean","std","max","min"] for w in SMP_ROLL_WINS]
    )
    m2 = m1.drop(columns=[c for c in smp_ar_cols if c in m1.columns])

    m1.to_parquet(OUT / "model1_train.parquet", index=False)
    m2.to_parquet(OUT / "model2_train.parquet", index=False)

    print(f"[Step 8] Model 1: {m1.shape}  →  output/model1_train.parquet")
    print(f"[Step 8] Model 2: {m2.shape}  →  output/model2_train.parquet")

    return m1, m2
