"""
기존 parquet 파일 → PostgreSQL 초기 적재 (1회 실행).

실행 순서:
    cd Powersignal
    python 02.database/load_initial.py

사전 조건:
    1. PostgreSQL 설치 & powersignal DB 생성 완료
    2. 프로젝트 루트 .env 에 PG_URL 설정
    3. schema.sql 실행 완료
       psql -U postgres -d powersignal -f 02.database/schema.sql
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd

# 프로젝트 루트를 sys.path 에 추가해 db 모듈을 임포트
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "02.database"))

from db import upsert, get_engine, rename_for_db  # noqa: E402

PROC = ROOT / "00.collector" / "data" / "processed"
FEAT = ROOT / "00.collector" / "data" / "features"
PREPROC_OUT = ROOT / "01.preprocessing" / "output"
MANUAL = ROOT / "00.collector" / "data" / "manual"

engine = get_engine()


def _load_parquets(folder: Path) -> pd.DataFrame:
    files = sorted(folder.glob("*.parquet"))
    if not files:
        raise FileNotFoundError(f"parquet 없음: {folder}")
    return pd.concat([pd.read_parquet(f) for f in files], ignore_index=True)


# ─────────────────────────────────────────────────────────────────────
# 1. smp_dayahead
# ─────────────────────────────────────────────────────────────────────
def load_smp_dayahead():
    print("\n[1/6] smp_dayahead 적재 중...")
    df = _load_parquets(PROC / "smp_dayahead")

    # datetime 생성: date(YYYYMMDD) + hour(1~24)
    base = pd.to_datetime(df["date"].astype(str), format="%Y%m%d", errors="coerce")
    h = pd.to_numeric(df["hour"], errors="coerce")
    df["datetime"] = base + pd.to_timedelta(h, unit="h")
    df = df.rename(columns={"areaName": "area_name"})
    df = df[["datetime", "area_name", "smp", "jlfd", "slfd", "mlfd"]].dropna(subset=["datetime"])

    for col in ["smp", "jlfd", "slfd", "mlfd"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.drop_duplicates(["datetime", "area_name"])
    n = upsert(df, "smp_dayahead", ["datetime", "area_name"], engine)
    print(f"  → {n:,}행 완료  ({df['datetime'].min()} ~ {df['datetime'].max()})")


# ─────────────────────────────────────────────────────────────────────
# 2. gen_by_source_hist
# ─────────────────────────────────────────────────────────────────────
_FUEL_RENAME = {
    "원자력":      "gen_nuclear",
    "LNG":         "gen_lng",
    "유연탄":      "gen_bituminous",
    "무연탄":      "gen_anthracite",
    "신재생·기타": "gen_renewable",
    "수력":        "gen_hydro",
    "양수":        "gen_pumped",
    "유전":        "gen_oil",
}


def load_gen_by_source():
    print("\n[2/6] gen_by_source_hist 적재 중...")
    df = _load_parquets(PROC / "gen_by_source_hist")

    base = pd.to_datetime(df["tradeYmd"].astype(str), format="%Y%m%d", errors="coerce")
    h = pd.to_numeric(df["tradeNo"], errors="coerce")
    df["datetime"] = base + pd.to_timedelta(h, unit="h")
    df["amgo"] = pd.to_numeric(df["amgo"], errors="coerce")

    # long → wide pivot
    wide = (
        df.pivot_table(index="datetime", columns="fuelTpCd", values="amgo", aggfunc="sum")
        .rename(columns=_FUEL_RENAME)
    )
    wide.columns.name = None

    gen_cols = [c for c in wide.columns if c.startswith("gen_")]
    wide["gen_total"] = wide[gen_cols].sum(axis=1)
    for col in gen_cols:
        wide[f"{col}_ratio"] = wide[col] / wide["gen_total"].replace(0, np.nan)

    wide = wide.reset_index().drop_duplicates("datetime")
    n = upsert(wide, "gen_by_source_hist", ["datetime"], engine)
    print(f"  → {n:,}행 완료  ({wide['datetime'].min()} ~ {wide['datetime'].max()})")


# ─────────────────────────────────────────────────────────────────────
# 3. monthly_fuel_cost
# ─────────────────────────────────────────────────────────────────────
def load_monthly_fuel_cost():
    print("\n[3/6] monthly_fuel_cost 적재 중...")
    df = _load_parquets(PROC / "monthly_fuel_cost")
    df["month"] = pd.to_datetime(df["day"].astype(str), format="%Y%m", errors="coerce")
    df["untpc"] = pd.to_numeric(df["untpc"], errors="coerce")
    df = df.rename(columns={"fuelType": "fuel_type", "untpc": "cost"})
    df = df[["month", "fuel_type", "cost"]].drop_duplicates(["month", "fuel_type"])
    n = upsert(df, "monthly_fuel_cost", ["month", "fuel_type"], engine)
    print(f"  → {n:,}행 완료")


# ─────────────────────────────────────────────────────────────────────
# 4. smp_decision_count
# ─────────────────────────────────────────────────────────────────────
def load_smp_decision_count():
    print("\n[4/6] smp_decision_count 적재 중...")
    df = _load_parquets(PROC / "smp_decision_count")
    df["trade_date"] = pd.to_datetime(df["tradeDay"].astype(str), format="%Y%m%d", errors="coerce")
    df["cnt"] = pd.to_numeric(df["cnt"], errors="coerce")
    df = df.rename(columns={"fuelType": "fuel_type", "areaNm": "area_name"})
    df = df[["trade_date", "fuel_type", "area_name", "cnt"]].drop_duplicates(
        ["trade_date", "fuel_type", "area_name"]
    )
    n = upsert(df, "smp_decision_count", ["trade_date", "fuel_type", "area_name"], engine)
    print(f"  → {n:,}행 완료")


# ─────────────────────────────────────────────────────────────────────
# 5. asos_hourly  (대용량 — 파일 단위로 처리)
# ─────────────────────────────────────────────────────────────────────
_ASOS_RENAME = {"ta": "temp_c", "hm": "humidity_pct", "ws": "wind_speed_ms", "td": "dew_point_c"}


def load_asos_hourly():
    print("\n[5/6] asos_hourly 적재 중... (파일이 많아 시간이 걸립니다)")
    folder = PROC / "asos_hourly"
    files = sorted(folder.glob("*.parquet"))
    total = 0
    for i, f in enumerate(files, 1):
        df = pd.read_parquet(f).rename(columns=_ASOS_RENAME)
        df["datetime"] = pd.to_datetime(df["tm"], format="%Y-%m-%d %H:%M", errors="coerce")
        df = df.rename(columns={"stn": "stn_id", "stnId": "stn_id"})
        cols = ["datetime", "stn_id", "temp_c", "humidity_pct", "wind_speed_ms", "dew_point_c"]
        df = df[[c for c in cols if c in df.columns]].dropna(subset=["datetime"])
        for col in ["temp_c", "humidity_pct", "wind_speed_ms", "dew_point_c"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        n = upsert(df, "asos_hourly", ["datetime", "stn_id"], engine)
        total += n
        if i % 20 == 0:
            print(f"  {i}/{len(files)} 파일 처리... 누적 {total:,}행")
    print(f"  → 총 {total:,}행 완료")


# ─────────────────────────────────────────────────────────────────────
# 6. nbtp_monthly (수요자원거래시장 NBTP, 수동 Excel)
# ─────────────────────────────────────────────────────────────────────
def load_nbtp_monthly():
    print("\n[6/7] nbtp_monthly 적재 중...")
    import openpyxl

    xlsx = MANUAL / "2026년도 3월 수요자원거래시장 현황 상세.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[wb.sheetnames[3]]  # 낙찰결과표 현황

    # 낙찰NBTP 행 위치 (연도 블록별)
    nbtp_accepted_rows = [5, 12, 19, 26, 33]
    records = []
    for nbtp_row in nbtp_accepted_rows:
        for col in range(3, 15):  # C~N = 1월~12월
            year  = ws.cell(row=nbtp_row - 2, column=col).value
            month = ws.cell(row=nbtp_row - 1, column=col).value
            if not isinstance(year, (int, float)):
                continue
            v_acc = ws.cell(row=nbtp_row,     column=col).value
            v_bid = ws.cell(row=nbtp_row + 1, column=col).value
            records.append({
                "month":         pd.Timestamp(int(year), int(month), 1),
                "nbtp_accepted": v_acc if v_acc not in (None, "-") else None,
                "nbtp_bid":      v_bid if v_bid not in (None, "-") else None,
            })

    df = pd.DataFrame(records)
    df = df.dropna(subset=["nbtp_accepted"]).reset_index(drop=True)
    n = upsert(df, "nbtp_monthly", ["month"], engine)
    print(f"  → {n:,}행 완료  ({df['month'].min().date()} ~ {df['month'].max().date()})")


# ─────────────────────────────────────────────────────────────────────
# 7. model_features (전처리 완료 parquet)
# ─────────────────────────────────────────────────────────────────────
def load_model_features():
    print("\n[7/7] model_features 적재 중...")

    for model_id, fname in [("model1", "model1_train.parquet"), ("model2", "model2_train.parquet")]:
        path = PREPROC_OUT / fname
        if not path.exists():
            print(f"  [{model_id}] {path} 없음 — 건너뜀")
            continue

        df = pd.read_parquet(path)
        df = rename_for_db(df)   # 한글 컬럼명 → 영문
        df["model_id"] = model_id
        df["datetime"] = pd.to_datetime(df["datetime"])

        # model2 는 SMP 자기회귀 컬럼 없음 → 스키마와 맞게 NaN 컬럼 추가
        smp_ar_cols = (
            [f"smp_lag{l}" for l in [1, 24, 48, 72, 168, 336]]
            + [f"smp_roll_{s}_{w}" for s in ["mean","std","max","min"] for w in [24, 168]]
        )
        for col in smp_ar_cols:
            if col not in df.columns:
                df[col] = np.nan

        n = upsert(df, "model_features", ["datetime", "model_id"], engine)
        print(f"  [{model_id}] → {n:,}행 완료  ({df['datetime'].min()} ~ {df['datetime'].max()})")


# ─────────────────────────────────────────────────────────────────────
# 실행
# ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="기존 parquet → PostgreSQL 초기 적재")
    parser.add_argument(
        "--tables", default="all",
        help="쉼표 구분 테이블명 또는 'all'. "
             "예: smp_dayahead,asos_hourly,model_features"
    )
    args = parser.parse_args()

    targets = (
        ["smp_dayahead", "gen_by_source_hist", "monthly_fuel_cost",
         "smp_decision_count", "asos_hourly", "nbtp_monthly", "model_features"]
        if args.tables == "all"
        else [t.strip() for t in args.tables.split(",")]
    )

    fn_map = {
        "smp_dayahead":       load_smp_dayahead,
        "gen_by_source_hist": load_gen_by_source,
        "monthly_fuel_cost":  load_monthly_fuel_cost,
        "smp_decision_count": load_smp_decision_count,
        "asos_hourly":        load_asos_hourly,
        "nbtp_monthly":       load_nbtp_monthly,
        "model_features":     load_model_features,
    }

    print("=" * 60)
    print("Powersignal DB 초기 적재 시작")
    print(f"대상: {', '.join(targets)}")
    print("=" * 60)

    for t in targets:
        if t in fn_map:
            fn_map[t]()
        else:
            print(f"[경고] 알 수 없는 테이블: {t}")

    print("\n" + "=" * 60)
    print("초기 적재 완료")
    print("=" * 60)
